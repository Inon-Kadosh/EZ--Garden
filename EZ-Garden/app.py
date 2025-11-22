import firebase_admin
import requests
import asyncio
import json
import openpyxl
import os
import tempfile
from io import BytesIO
from fpdf import FPDF
from datetime import datetime, timedelta
from flask_mail import Mail, Message
from firebase_admin import credentials, firestore
from flask import Flask, render_template, request, redirect, session, send_from_directory,Response,jsonify,url_for
from flask_caching import Cache
from openplantbook_sdk import OpenPlantBookApi, MissingClientIdOrSecret, ValidationError
import pandas as pd




SENSOR_API_LOGIN_URL = "https://atapi.atomation.net/auth/login"
SENSOR_API_READINGS_URL = "https://atapi.atomation.net/sensors_readings"
OPENPLANT_API_BASE_URL = "https://open.plantbook.io/api/v1"
SENSOR_APP_VERSION = "1.0.0"
SENSOR_ACCESS_TYPE = 5



BASE_URL = "https://open.plantbook.io/api/v1"
CLIENT_ID = "tC0yqwKuCLpFRkFNjTFsHugNJo6poO0I8neJFZuR"
CLIENT_SECRET = "Fny45ACTmtXTOwpC5j3b00HeDgvboHZMh8mjbH190FmeNBNsAWH3mjNm6gYNg38FMyszvzTzhsy2GRdhPb8YcY6uMmD4NYnhhUKPac8a6h1Zgh35IIjDrfRDqYiGmTdT"



app = Flask(__name__)



app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'ezgarden.sce@gmail.com'  
app.config['MAIL_PASSWORD'] = 'pyas pbfo xstu hksu'  
app.config['MAIL_DEFAULT_SENDER'] = 'ezgarden.sce@gmail.com'

mail = Mail(app)


@app.route('/send-email', methods=['POST'])
def send_email():
    try:
        # קבלת האימייל מהבקשה
        data = request.get_json()
        email = data.get('email')
        if not email:
            return jsonify({'error': 'Email is required'}), 400

        print("Creating email...")
        # יצירת מייל
        msg = Message('Graph Data Notification', sender=app.config['MAIL_USERNAME'], recipients=[email])
        msg.body = """
        Dear User,

        The data you requested is now available. Please visit the application to view or download the data.

        Best regards,
        Your Team
        """

        print("Sending email...")
        # שליחת המייל
        mail.send(msg)
        print("Email sent successfully.")

        return jsonify({'message': f'Notification sent successfully to {email}'}), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500






app.config['CACHE_TYPE'] = 'SimpleCache'
app.config['CACHE_DEFAULT_TIMEOUT'] = 3600  
cache = Cache(app)



app.secret_key = "ezgarden"



cred = credentials.Certificate("ez-garden-firebase-adminsdk-rr8ir-aad6b3e5d2.json")
firebase_admin.initialize_app(cred)
db = firestore.client()



@app.context_processor
def inject_user_info():
    """Inject user info into all templates."""
    user_email = session.get("user_email")
    
    return {"user_email": user_email}



@app.route('/')
def home():
    return render_template('home1.html')


#comment 

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        if not email or not password:
            return render_template('login.html', error_message="Both email and password are required!")

        employees_ref = db.collection("employees").where("company_email", "==", email).where("password", "==", password).get()
        # if employee
        if employees_ref:
            employee_data = employees_ref[0].to_dict()
            role = employee_data.get("role", 1)  
            company_name = employee_data.get("company_name")

            #Session
            session['user_email'] = email
            session['user_role'] = role
            session['company_name'] = company_name

            if role == 1:  
                return redirect("/home_company_1")
            elif role == 2:   
                return redirect("/home_company_2")
            elif role == 3:  
                return redirect("/home_company_3")

        # if user
        users_ref = db.collection("users").where("email", "==", email).where("password", "==", password).get()
        if users_ref:
            session['user_email'] = email
            session['user_type'] = "private"
            return redirect("/home3")

        return render_template('login.html', error_message="Invalid email or password!")

    return render_template('login.html')



@app.route('/register_type')
def register_type():
    return render_template('register_type.html')



@app.route('/register_private', methods=['GET', 'POST'])
def register_private():
    if request.method == 'POST':
        fullname = request.form.get('fullname')
        email = request.form.get('email')
        password = request.form.get('password')

        if not fullname or not email or not password:
            return render_template('register_private.html', error_message="All fields are required!")

        #Saving data in Firebase
        user_data = {
            "type": "private",
            "fullname": fullname,
            "email": email,
            "password": password
        }
        db.collection("users").add(user_data)
        
        try:
            msg = Message('Welcome to EZ-GARDEN!',
                      recipients=[email])
            msg.body = 'Thank you for registering! We are excited to have you on board.'
            mail.send(msg)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Failed to send email: {e}")

        # success
        return render_template('success.html', message=f"User {fullname} successfully registered with email {email}!")
    
    

    return render_template('register_private.html')



@app.route('/register_company', methods=['GET', 'POST'])
def register_company():
    if request.method == 'POST':
        company_name = request.form.get('company_name')
        name = request.form.get('name')
        company_email = request.form.get('email')
        password = request.form.get('password')

        # Checking if all fields are complete
        if not company_name or not name or not company_email or not password:
            return render_template('register_company.html', error_message="All fields are required!")

        #Saving data in Firebase
        company_data = {
            "company_name": company_name,
            "name": name,
            "company_email": company_email,
            "password": password,
            "role": 3  
        }
        db.collection("employees").add(company_data)
        
        try:
            msg = Message('Welcome to EZ-GARDEN!',
                      recipients=[company_email])
            msg.body = 'Thank you for registering! We are excited to have you on board.'
            mail.send(msg)
            print("Email sent successfully!")
        except Exception as e:
            print(f"Failed to send email: {e}")

        # success 
        success_message = f"The company {company_name} has been successfully registered with the email {company_email}!"
        return render_template('success.html', message=success_message)

    return render_template('register_company.html')



@app.route('/home_company_3')
def home_company_3():
    return render_template('home_company_3.html')

@app.route('/data_for_kids')
def data_for_kids():
    return render_template('data_for_kids.html')

@app.route('/planting_experiment')
def planting_experiment():
    return render_template('planting_experiment.html')

@app.route('/saving_water_tips')
def saving_water_tips():
    return render_template('saving_water_tips.html')



neighborhoods = [
    "NEVE NOY", "NAHAL BEKA", "OLD CITY", "PELAH 7", "NEVE ZEEV",
    "ALEF", "TET", "NEOT LON", "HEY", "BET", "VAV", "NAHAL ASHAN",
    "KALANIYOT", "D", "GIMEL", "RAMOT", "YUD ALEF"
]

@app.route('/home_company_2', methods=['GET', 'POST'])
def home_company_2():
    if request.method == 'POST':
        # Extract data from the form
        date = request.form.get('date')
        time = request.form.get('time')
        area = request.form.get('area')
        duration = request.form.get('duration')

        # Simulate saving the data (e.g., save to database)
        print(f"Manual watering settings received: Date={date}, Time={time}, Area={area}, Duration={duration}")

        # Return a JSON response (if using AJAX)
        return jsonify({"success": True, "message": "Settings saved successfully!"})

    # Render the HTML template for the page
    return render_template('home_company_2.html', neighborhoods=neighborhoods)



@app.route('/home_company_1')
def home_company_1():
    return render_template('home_company_1.html')



@app.route("/employee_management", methods=["GET", "POST"])
def employee_management():
    company_email = session.get("user_email")
    company_name = session.get("company_name")
    
    # Connected administrator email
    if not company_email:
        print("No company email in session. Redirecting to login.")
        return redirect('/login')  

    if request.method == 'POST':
        # get data
        employee_name = request.form.get('employee_name')
        employee_password = request.form.get('employee_password')
        employee_role = request.form.get('employee_role')

        # chack data
        if not employee_name or not employee_password or not employee_role:
            print("Missing field(s) in the form submission.")
            return render_template(
                'employee_management.html',
                error_message="All fields are required!",
                employees=get_employees_from_db(company_email)
            )

        if not employee_role.isdigit() or int(employee_role) not in [1, 2, 3]:
            print("Invalid role value.")
            return render_template(
                'employee_management.html',
                error_message="Role must be 1, 2, or 3!",
                employees=get_employees_from_db(company_email)
            )

        # Preparing data for Firebase
        employee_data = {
            "company_name": company_name.strip(),
            "company_email": company_email.strip(),
            "name": employee_name.strip(),
            "password": employee_password.strip(),
            "role": int(employee_role)
            
        }

        #Saving data in Firebase
        try:
            print("Attempting to add employee to Firebase:", employee_data)
            db.collection("employees").add(employee_data)
            print("Successfully added employee:", employee_data)
        except Exception as e:
            print("Error adding employee to Firebase:", e)
            return render_template(
                'employee_management.html',
                error_message="Failed to save employee to the database!",
                employees=get_employees_from_db(company_email)
            )

        return redirect('/employee_management')

    # Pulling data from employees with the same email
    try:
        employees = get_employees_from_db(company_email)
        print("Employees retrieved from database:", employees)
    except Exception as e:
        print("Error retrieving employees:", e)
        return render_template(
            'employee_management.html',
            error_message="Failed to retrieve employees!",
            employees=[]
        )

    return render_template('employee_management.html', employees=employees)



def get_employees_from_db(company_email):
    try:
        employees_ref = db.collection("employees").where("company_email", "==", company_email).stream()
        employees = []
        for emp in employees_ref:
            employee_data = emp.to_dict()
            employee_data['id'] = emp.id  # save id
            employees.append(employee_data)
        return employees
    except Exception as e:
        print("Error retrieving employees from Firebase:", e)
        return []



@app.route("/delete_employee/<employee_id>", methods=["POST"])
def delete_employee(employee_id):
    try:
        # Deleting an employee by id
        db.collection("employees").document(employee_id).delete()
        print(f"Employee {employee_id} deleted successfully.")
    except Exception as e:
        print(f"Error deleting employee {employee_id}: {e}")
        return redirect('/employee_management')

    # back
    return redirect('/employee_management')



@app.route('/map')
def map():
    return render_template('map.html')




@app.route('/weather', methods=['GET', 'POST'])
def weather():
    if request.method == 'POST':
        location = request.form.get('location')  # get location

        if not location:
            return "Please provide a location.", 400

        # call API
        url = f"https://api.tomorrow.io/v4/weather/realtime?location={location}&apikey=cAKQ5PlikazuWRCN01CpQnFuzes6rcmt"
        headers = {"accept": "application/json"}

        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()  
            data = response.json()

            return render_template("weather_data.html", weather_data=data)
        except requests.exceptions.RequestException as e:
            return f"Failed to fetch weather data: {e}", 500

    return render_template("weather.html")



@app.route('/static/map.geojson')
def geojson():
    return send_from_directory('static', 'map.geojson')



@app.route('/neighborhood/<name>')
def neighborhood(name):
    return f"<h1>Welcome to {name.capitalize()} neighborhood</h1>"



@app.route('/home3')
def home3():
    user_email = session.get('user_email')  # get email

    if not user_email:
        return redirect('/login') 

    #Searched on Firebase
    plants_ref = db.collection("plants")
    query = plants_ref.where("user_email", "==", user_email).get()

    #Create a plant list
    plants = [
    {
        "id": plant.id,  # Document ID
        "area": plant.to_dict().get("area", "Unknown"),  
        "type": plant.to_dict().get("type", "Unknown")   
    }
    for plant in query
    ]

    # sending home3.html
    return render_template('home3.html', plants=plants)




@app.route('/export_to_excel', methods=['GET'])
def export_to_excel():
    user_email = session.get('user_email')
    if not user_email:
        return redirect('/login')

    # חיפוש נתוני הצמחים ב-Firebase
    plants_ref = db.collection("plants")
    query = plants_ref.where("user_email", "==", user_email).get()

    # יצירת קובץ Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Plants"

    # הוספת כותרות
    sheet.append(["Area", "Type"])

    # הוספת נתונים
    for plant in query:
        plant_data = plant.to_dict()
        sheet.append([plant_data.get("area", "Unknown"), plant_data.get("type", "Unknown")])

    # שמירה ל-BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # החזרת קובץ למשתמש
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=plants.xlsx"}
    )


@app.route('/delete_plant/<plant_id>', methods=['POST'])
def delete_plant(plant_id):
    user_email = session.get('user_email')

    if not user_email:
        return redirect('/login')

    db.collection("plants").document(plant_id).delete()

    return redirect('/home3')



@app.route('/add_plant', methods=['GET', 'POST'])
def add_plant():
    user_email = session.get('user_email')  # get email

    if not user_email:
        return redirect('/login') 

    if request.method == 'POST':
        area = request.form.get('area')
        type = request.form.get('type')

        if not area or not type:
            return "<h1 style='text-align:center; color:red;'>All fields are required!</h1>"

        #Saving the plant in Firebase
        plant_data = {
            "area": area,
            "type": type,
            "user_email": user_email
        }
        db.collection("plants").add(plant_data)

        return redirect("/home3")

    return render_template('add_plant.html')



# Access Token OpenPlantBookApi
def get_access_token(client_id, client_secret):
    url = "https://open.plantbook.io/api/v1/token/"
    data = {
        "grant_type": "client_credentials",
        "client_id": 'tC0yqwKuCLpFRkFNjTFsHugNJo6poO0I8neJFZuR',
        "client_secret": 'Fny45ACTmtXTOwpC5j3b00HeDgvboHZMh8mjbH190FmeNBNsAWH3mjNm6gYNg38FMyszvzTzhsy2GRdhPb8YcY6uMmD4NYnhhUKPac8a6h1Zgh35IIjDrfRDqYiGmTdT',
        "scope": "read"
    }

    response = requests.post(url, data=data)

    if response.status_code == 200:
        token_data = response.json()
        return token_data.get("access_token")
    else:
        print(f"Error: Unable to fetch token. Status code: {response.status_code}")
        print(f"Response: {response.text}")
        return None



#Getting information for a plant
async def get_plant_details(plant_type, access_token):
    url = f"https://open.plantbook.io/api/v1/plant/search?alias={plant_type}"
    headers = {"Authorization": f"Bearer {access_token}"}

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            search_results = response.json()
            if search_results and 'results' in search_results and len(search_results['results']) > 0:
                plant = search_results['results'][0]  
                plant_pid = plant['pid']
                
             
                detail_url = f"https://open.plantbook.io/api/v1/plant/detail/{plant_pid}"
                detail_response = requests.get(detail_url, headers=headers)
                
                if detail_response.status_code == 200:
                    return detail_response.json()
                else:
                    print(f"Error fetching plant details: {detail_response.status_code}")
                    return None
            else:
                print("No results found for the plant type.")
                return None
        else:
            print(f"Error fetching plant data: {response.status_code}")
            return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None



def get_sensor_api_token():
    print('GET TOKEN_SENSOR')

    url = "https://atapi.atomation.net/api/v1/s2s/v1_0/auth/login"
    headers = {
        "accept": "application/json",
        "app_version": "v1.0",
        "access_type": "5",
        "Content-Type": "application/json"
    }
    payload = {
        "email": "sce@atomation.net",
        "password": "123456"
    }

    try:
        response = requests.post(url, headers=headers, data=json.dumps(payload))
        response.raise_for_status()
        
        data = response.json()
        token = data.get("data", {}).get("token")
        if token:
            print("Token successfully retrieved!")
            return token
        else:
            print("Error: Token not found in response.")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error fetching token: {e}")
        return None



def get_last_sensor_reading(token, mac_address):
 
    url = "https://atapi.atomation.net/api/v1/s2s/v1_0/sensors_readings"

    if not mac_address:
        print("MAC address is missing.")
        return {"error": "MAC address is missing."}

    #Create a date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(hours=1)
    formatted_start_date = start_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_end_date = end_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")

    print("Start Date:", formatted_start_date)
    print("End Date:", formatted_end_date)


    payload = {
        "filters": {
            "start_date": formatted_start_date,
            "end_date": formatted_end_date,
            "mac": [mac_address],
            "createdAt": True
        },
        "limit": {
            "page": 1,
            "page_size": 1
        }
    }


    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    try:
        # Sending to api
        response = requests.post(url, data=json.dumps(payload), headers=headers)
        print(f"Response Status Code: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            print("Response Data:", data)

            # Check if data exists
            if "data" in data and "readings_data" in data["data"]:
                readings_data = data["data"]["readings_data"]
                if isinstance(readings_data, list) and len(readings_data) > 0:
                    return readings_data[0]  # last reading
                else:
                    print("No data found in the sensor.")
                    return {"message": "No sensor data available."}
            else:
                print("Unexpected data structure from API response.")
                return {"error": "Unexpected data structure from API response."}
        else:
            print(f"Error receiving sensor reading: {response.status_code}")
            print(":", response.text)
            return {"error": f"Sensor API returned status code {response.status_code}"}
    except json.JSONDecodeError:
        print("Response is not a valid JSON")
        return {"error": "Response is not a valid JSON."}
    except Exception as e:
        print(f"General error: {e}")
        return {"error": f"An unexpected error occurred: {e}"}



#OpenPlantBookApi
@app.route('/plant_details/<plant_type>')
def plant_details(plant_type):

    #Access Token
    access_token = get_access_token(CLIENT_ID, CLIENT_SECRET)
    if not access_token:
        return "Error: Unable to retrieve access token.", 500

    details = asyncio.run(get_plant_details(plant_type, access_token))

    #chacking    
    if not details:
        return "Error: Unable to retrieve plant details.", 500
    
    
    
    #atomation
    sensor_token = cache.get("sensor_api_token")
    if not sensor_token:
        # if token empty
        sensor_token = get_sensor_api_token()
        if not sensor_token:
            return "Error: Unable to retrieve sensor token.", 500


        #save token for 1 hour
        cache.set("sensor_api_token", sensor_token, timeout=3600)

    #checking
    sensor_cache_key = "last_sensor_reading"
    sensor_data = cache.get(sensor_cache_key)
    if not sensor_data:
        #get sensor data
        sensor_data = get_last_sensor_reading(sensor_token, "E9:19:79:09:A1:AD")
        if sensor_data:
            cache.set(sensor_cache_key, sensor_data, timeout=60)  # save the last reading for 1 min

    return render_template('plant_details.html', plant=details, sensor=sensor_data)



@app.route('/plants_by_area', methods=['GET'])
def plants_by_area():
    neighborhood = request.args.get('neighborhood')  #  get neighborhood name

    if not neighborhood:
        return "<h1 style='color:red;'>Neighborhood not specified!</h1>"

    user_email = session.get('user_email')  # get email

    if not user_email:
        return redirect('/login')  

    # Search Firebase by neighborhood
    plants_ref = db.collection("plants")
    query = (
        plants_ref
        .where("user_email", "==", user_email)
        .where("area", "==", neighborhood)
        .get()
    )

    # Create a list of plants
    plants = [
        {
            "id": plant.id,  # Document ID
            "area": plant.to_dict().get("area", "Unknown"),  
            "type": plant.to_dict().get("type", "Unknown")   
        }
        for plant in query
    ]

    return render_template('plants_by_area.html', plants=plants, neighborhood=neighborhood)




# פונקציה לטעינת הקובץ maintenance.json
def load_data():
    return pd.read_json("static/maintenance.json")

# נתיב לעמוד "דף תקלות"
@app.route('/issues')
def issues_dashboard():
    issues = [
        {
            "title": "Sensor Malfunction",
            "description": "The sensor is not responding to requests. Please check the wiring.",
            "image": "/static/images/sensor_error.png"
        },
        {
            "title": "Overheating",
            "description": "The device is overheating and needs immediate attention.",
            "image": "/static/images/overheating.png"
        },
        {
            "title": "Network Issue",
            "description": "The system is unable to connect to the network. Check the configuration.",
            "image": "/static/images/network_error.png"
        }
    ]
    return render_template('issues.html', issues=issues)

# פונקציה לקריאת נתונים מקובץ JSON
def load_alerts():
    with open('static/alerts.json', 'r', encoding='utf-8') as file:
        return json.load(file)

@app.route('/alerts')
def alerts():
    alerts_data = load_alerts()
    return render_template('alerts.html', alerts=alerts_data)


# נתיב לעמוד "תחזוקה תקופתית"
@app.route('/maintenance')
def maintenance():
    data = load_data()
    return render_template(
        "maintenance.html",
        months=data["Month"].tolist(),
        temperatures=data["Temperature (°C)"].tolist(),
        humidity=data["Humidity (%)"].tolist(),
        temperature_data=data[["Month", "Start Date", "End Date", "Temperature (°C)"]].to_dict(orient="records"),
        humidity_data=data[["Month", "Start Date", "End Date", "Humidity (%)"]].to_dict(orient="records")
    )
#פונקציה שמחזירה את נתוני היעילות בפורמט JSON
@app.route('/get_efficiency_data', methods=['GET'])
def get_efficiency_data():
    with open('static/efficiency_data.json', 'r', encoding='utf-8') as file:
        data = json.load(file)
    return jsonify(data)
#גרפים עבור יעלות מערכת 
@app.route('/efficiency_dashboard', methods=['GET'])
def efficiency_dashboard():
    return render_template('efficiency_dashboard.html')

@app.route('/education')
def education():
    # Load schools data from JSON
    with open('static/schools.json', 'r', encoding='utf-8') as file:
        schools = json.load(file)

    # Find the school with the highest water consumption
    max_consumption_school = max(schools, key=lambda school: school['water_consumption'])

    # Render the template and pass the data
    return render_template('education.html', schools=schools, max_school=max_consumption_school)










with open("static/water_consumption.json", "r") as file:
    data_water = json.load(file)
    
    
@app.route("/data_water")
def get_water():
    print("Loading JSON file...")
    with open("static/water_consumption.json", "r") as file:
        data_water = json.load(file)
    print("JSON loaded successfully!")

    # Retrieve selected dates from the session
    start_date_str = session.get("start_date")
    end_date_str = session.get("end_date")
    print("Session Dates:", start_date_str, "to", end_date_str)

    if not start_date_str or not end_date_str:
        return jsonify({"error": "Date range not selected"}), 400

    # Convert dates to datetime objects
    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
    except ValueError as e:
        print("Date parsing error:", e)
        return jsonify({"error": "Invalid date format in session"}), 400

    print("Filtering data...")
    filtered_data = {
        "dates": [],
        "before_installation": [],
        "after_installation": []
    }

    for idx, date_str in enumerate(data_water["dates"]):
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            if start_date <= date_obj <= end_date:
                filtered_data["dates"].append(date_str)
                filtered_data["before_installation"].append(data_water["before_installation"][idx])
                filtered_data["after_installation"].append(data_water["after_installation"][idx])
        except ValueError as e:
            print(f"Invalid date at index {idx}: {date_str}, Error: {e}")
            continue

    print("Filtered Data:", filtered_data)

    if not filtered_data["dates"]:
        return jsonify({"error": "No data found in the selected date range"}), 404

    return jsonify(filtered_data)


def get_sensor_by_date(token, mac_address, start_date, end_date):
    print("TAKE_DATA_SENSOR")
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(hours=24)
    formatted_start_date = start_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_end_date = end_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    url = "https://atapi.atomation.net/api/v1/s2s/v1_0/sensors_readings"

    if not mac_address:
        print("MAC address is missing.")
        return {"error": "MAC address is missing."}

    # Format start and end dates
    formatted_start_date = start_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_end_date = end_date.strftime("%Y-%m-%dT%H:%M:%S.%fZ")

    print("Start Date:", formatted_start_date)
    print("End Date:", formatted_end_date)

    payload = {
        "filters": {
            "start_date": formatted_start_date,
            "end_date": formatted_end_date,
            "mac": [mac_address],
            "sample-time": True
        },
        "limit": {
            "page": 1,
            "page_size": 1000
        }
    }

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    try:
        # Sending request to API
        response = requests.post(url, data=json.dumps(payload), headers=headers)
        print(f"Response Status Code: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            print("Response Data (Full):", data)

            # Validate structure
            if "data" in data and "readings_data" in data["data"]:
                readings_data = data["data"]["readings_data"]
                print(f"Number of readings: {len(readings_data)}")
                if isinstance(readings_data, list) and len(readings_data) > 0:
                    return readings_data  # Return all readings
                else:
                    print("No data found in the sensor.")
                    return {"message": "No sensor data available."}
            else:
                print("Unexpected data structure from API response.")
                return {"error": "Unexpected data structure from API response."}
        else:
            print(f"Error receiving sensor reading: {response.status_code}")
            print("Error Response Text:", response.text)
            return {"error": f"Sensor API returned status code {response.status_code}"}
    except json.JSONDecodeError:
        print("Response is not a valid JSON")
        return {"error": "Response is not a valid JSON."}
    except Exception as e:
        print(f"General error: {e}")
        return {"error": f"An unexpected error occurred: {e}"}
    

@app.route("/graphs_select_date", methods=["GET"])
def graphs_select_date():

    return render_template("graphs_select_date.html")


@app.route("/data_sensor_by_date")
def data_sensor_by_date():
    sensor_token = cache.get("sensor_api_token")
    if not sensor_token:
        sensor_token = get_sensor_api_token()
        if not sensor_token:
            return jsonify({"error": "Unable to retrieve sensor token"}), 500
        cache.set("sensor_api_token", sensor_token, timeout=3600)

    sensor_cache_key = "last_sensor_reading"
    sensor_data = cache.get(sensor_cache_key)
    if not sensor_data:
        # ----> קבלת התאריכים מ-Session <----
        user_start_str = session.get("start_date")  # "YYYY-MM-DD"
        user_end_str = session.get("end_date")  # "YYYY-MM-DD"

        if not user_start_str or not user_end_str:
            return jsonify({"error": "Date range not provided"}), 400

        # המרת תאריכים ל-datetime
        try:
            start_date = datetime.strptime(user_start_str, "%Y-%m-%d")
            end_date = datetime.strptime(user_end_str, "%Y-%m-%d")
            # הוספת סוף היום לתאריך הסיום
            end_date += timedelta(hours=23, minutes=59, seconds=59)
        except ValueError as e:
            return jsonify({"error": f"Invalid date format: {e}"}), 400

        print(f"Requesting sensor data for range: {start_date} -> {end_date}")

        try:
            sensor_data = get_sensor_by_date(sensor_token, "E9:19:79:09:A1:AD", start_date, end_date)
            if sensor_data:
                cache.set(sensor_cache_key, sensor_data, timeout=60)
            else:
                return jsonify({"error": "No sensor data available"}), 404
        except Exception as e:
            return jsonify({"error": f"Error retrieving sensor data: {e}"}), 500

    try:
        filtered_data = [
            {
                "Temperature": reading["Temperature"],
                "Humidity": reading["Humidity"],
                "SampleTime": reading["sample_time_utc"]
            }
            for reading in sensor_data
        ]
    except KeyError as e:
        return jsonify({"error": f"Missing key: {e}"}), 500

    print("Final Filtered Data:", filtered_data)
    session["final_data_json"] = json.dumps(filtered_data)

    return jsonify(filtered_data)



@app.route('/graphs', methods=['GET', 'POST'])
def graphs():
    # מוציאים את התאריכים מתוך ה-Session
    start_date_str = session.get("start_date")
    end_date_str   = session.get("end_date")

    if not start_date_str or not end_date_str:
        return "<h3>No dates in session. Please go back and select dates.</h3>"

    # המרתם ל-datetime (לא חובה, אלא אם נרצה לעבד אותם בפייתון)
    start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
    end_dt   = datetime.strptime(end_date_str, "%Y-%m-%d")
    

    return render_template("graphs.html",
                           start_date=start_dt,
                           end_date=end_dt)


@app.route("/handle_dates", methods=["POST"])
def handle_dates():
    # מקבלים את התאריכים מהטופס (POST)
    start_date_str = request.form.get("startDate")  # "YYYY-MM-DD"
    end_date_str   = request.form.get("endDate")    # "YYYY-MM-DD"

    # שומרים אותם ב-Session
    session["start_date"] = start_date_str
    session["end_date"]   = end_date_str

    print(f"Dates saved in session: {start_date_str} -> {end_date_str}")

    # מפנים לנתיב שמציג את הגרפים
    return redirect(url_for("graphs"))


@app.route('/export_graph', methods=['POST'])
def export_graph():
    try:
        # קבלת הנתונים מהבקשה
        data = request.get_json()
        if not data or "graph_data" not in data:
            print("Invalid request data received")  # Log שגיאה
            return jsonify({"error": "Invalid request data"}), 400

        graph_data = data["graph_data"]
        labels = graph_data.get("labels", [])
        datasets = graph_data.get("datasets", [])

        # לוגים של הנתונים שהתקבלו
        print(f"Labels received: {labels}")
        for dataset in datasets:
            print(f"Dataset '{dataset['label']}' received with {len(dataset['data'])} points")

        # יצירת קובץ Excel
        output = BytesIO()
        df = pd.DataFrame({"Labels": labels})
        for dataset in datasets:
            df[dataset["label"]] = dataset["data"]

        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        print("Excel file created successfully")  # Log להצלחה

        # שליחת הקובץ להורדה
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=water_consumption_data.xlsx"}
        )
    except Exception as e:
        print(f"Error exporting graph: {e}")  # Log לשגיאה
        return jsonify({"error": str(e)}), 500


@app.route('/send-email-with-files', methods=['POST'])
def send_email_with_files():
    try:
        data = request.get_json()
        email = data.get('email')
        if not email:
            return jsonify({'error': 'Email is required'}), 400

        # יצירת קבצים
        pdf_path = os.path.join(tempfile.gettempdir(), 'graphs.pdf')
        create_pdf(pdf_path)

        excel_path = os.path.join(tempfile.gettempdir(), 'temp_humidity_data.xlsx')
        create_excel(excel_path)

        # יצירת מייל
        msg = Message(
            subject='Graph Data and Reports',
            sender=app.config['MAIL_USERNAME'],
            recipients=[email]
        )
        msg.body = """
        Dear User,

        Attached you will find the requested graphs and reports.

        Best regards,
        Your Team
        """

        # הוספת קבצים מצורפים
        with open(pdf_path, 'rb') as pdf_file:
            msg.attach('graphs.pdf', 'application/pdf', pdf_file.read())
        
        with open(excel_path, 'rb') as excel_file:
            msg.attach('temp_humidity_data.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', excel_file.read())

        mail.send(msg)
        print("Email with files sent successfully.")

        return jsonify({'message': f'Email with files sent successfully to {email}'}), 200

    except Exception as e:
        print(f"An error occurred: {e}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500


def create_pdf(pdf_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Example PDF Content", ln=True, align='C')
    pdf.output(pdf_path)
    print("PDF created successfully at", pdf_path)


def create_excel(excel_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Example Sheet"
    sheet.append(["Column1", "Column2", "Column3"])
    sheet.append(["Data1", "Data2", "Data3"])
    workbook.save(excel_path)
    print("Excel created successfully at", excel_path)




if __name__ == "__main__":
    app.run(debug=True)
    